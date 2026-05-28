import io
import json
import re
import zipfile
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import SessionLocal, criar_tabelas
from app.models.company_model import EmpresaModel
from app.models.product_model import ProdutoModel

EMPRESA_PADRAO_ID = 1
EMPRESA_PADRAO_NOME = "Empresa Local"

BASE_DIR = Path(__file__).resolve().parent.parent.parent
DATA_DIR = BASE_DIR / "data"
JSON_ANTIGO = DATA_DIR / "produtos.json"


def _abrir_xlsx(caminho_arquivo):
    """
    Abre um xlsx com openpyxl, corrigindo automaticamente o bug de
    activePane inválido comum nos exports da Shopee.
    """
    try:
        return openpyxl.load_workbook(caminho_arquivo, data_only=True)
    except Exception:
        validos = ("bottomRight", "bottomLeft", "topRight", "topLeft")
        with zipfile.ZipFile(caminho_arquivo, "r") as origem:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dest:
                for item in origem.infolist():
                    conteudo = origem.read(item.filename)
                    if item.filename.endswith(".xml"):
                        txt = conteudo.decode("utf-8", errors="ignore")
                        txt = re.sub(
                            r'activePane="([^"]*)"',
                            lambda m: m.group(0) if m.group(1) in validos else "",
                            txt,
                        )
                        conteudo = txt.encode("utf-8")
                    dest.writestr(item, conteudo)
            buf.seek(0)
        return openpyxl.load_workbook(buf, data_only=True)


def _xlsx_para_lista(caminho_arquivo, linha_cabecalho=0):
    """
    Lê um xlsx e devolve lista de dicts {coluna: valor}.
    Substitui pd.read_excel sem depender de pandas/numpy.
    """
    wb = _abrir_xlsx(caminho_arquivo)
    ws = wb.active
    todas = list(ws.iter_rows(values_only=True))
    wb.close()

    if not todas or linha_cabecalho >= len(todas):
        return []

    cab = [
        str(h).strip() if h is not None else f"_col{i}"
        for i, h in enumerate(todas[linha_cabecalho])
    ]
    return [{c: v for c, v in zip(cab, row)} for row in todas[linha_cabecalho + 1 :]]


def normalizar_sku(sku):
    return str(sku or "").strip().upper()


def normalizar_codigo(codigo):
    """Normaliza um código de marketplace (ID do produto)."""
    return str(codigo or "").strip().upper()


def codigos_para_texto(codigos):
    """Aceita lista ou string e devolve string normalizada separada por vírgula."""
    if not codigos:
        return ""
    if isinstance(codigos, str):
        partes = codigos.replace(";", ",").split(",")
    else:
        partes = list(codigos)
    limpos = []
    for c in partes:
        c = normalizar_codigo(c)
        if c and c not in limpos:
            limpos.append(c)
    return ",".join(limpos)


def texto_para_codigos(texto):
    if not texto:
        return []
    return [c for c in (normalizar_codigo(p) for p in str(texto).split(",")) if c]


def componentes_para_json(componentes):
    if not componentes:
        return "[]"

    componentes_limpos = []

    for componente in componentes:
        sku = normalizar_sku(componente.get("sku"))
        quantidade = float(componente.get("quantidade") or 1)

        if sku:
            componentes_limpos.append(
                {
                    "sku": sku,
                    "quantidade": quantidade,
                }
            )

    return json.dumps(componentes_limpos, ensure_ascii=False)


def json_para_componentes(texto):
    if not texto:
        return []

    try:
        dados = json.loads(texto)

        if isinstance(dados, list):
            return dados

        return []

    except Exception:
        return []


def garantir_empresa_padrao(db):
    empresa = (
        db.query(EmpresaModel).filter(EmpresaModel.id == EMPRESA_PADRAO_ID).first()
    )

    if empresa:
        return empresa

    empresa = EmpresaModel(
        id=EMPRESA_PADRAO_ID,
        nome=EMPRESA_PADRAO_NOME,
        documento="",
        plano="local",
        ativo=True,
    )

    db.add(empresa)
    db.commit()
    db.refresh(empresa)

    return empresa


def produto_para_dict(produto):
    return {
        "id": produto.id,
        "empresa_id": produto.empresa_id,
        "sku": produto.sku,
        "nome": produto.nome,
        "custo": produto.custo,
        "imposto": produto.imposto,
        "tipo": produto.tipo,
        "frete_gratis": produto.frete_gratis,
        "observacao": produto.observacao,
        "componentes": json_para_componentes(produto.componentes),
        "codigosExternos": getattr(produto, "codigos_externos", "") or "",
        "codigos_externos_lista": texto_para_codigos(
            getattr(produto, "codigos_externos", "")
        ),
        "created_at": produto.created_at.isoformat() if produto.created_at else None,
        "updated_at": produto.updated_at.isoformat() if produto.updated_at else None,
    }


def calcular_custo_kit(db, componentes, empresa_id):
    custo_total = 0

    for componente in componentes:
        sku = normalizar_sku(componente.get("sku"))
        quantidade = float(componente.get("quantidade") or 1)

        produto_componente = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.sku == sku)
            .first()
        )

        if not produto_componente:
            raise ValueError(f"Componente {sku} não encontrado.")

        custo_total += float(produto_componente.custo or 0) * quantidade

    return custo_total


def migrar_json_antigo_para_banco(db):
    if not JSON_ANTIGO.exists():
        return

    total_produtos = (
        db.query(ProdutoModel)
        .filter(ProdutoModel.empresa_id == EMPRESA_PADRAO_ID)
        .count()
    )

    if total_produtos > 0:
        return

    try:
        with open(JSON_ANTIGO, "r", encoding="utf-8") as arquivo:
            produtos_antigos = json.load(arquivo)
    except Exception:
        return

    if not isinstance(produtos_antigos, list):
        return

    for produto in produtos_antigos:
        sku = normalizar_sku(produto.get("sku"))

        if not sku:
            continue

        produto_model = ProdutoModel(
            empresa_id=EMPRESA_PADRAO_ID,
            sku=sku,
            nome=produto.get("nome") or sku,
            custo=float(produto.get("custo") or 0),
            imposto=float(produto.get("imposto") or 0),
            tipo=produto.get("tipo") or "unitario",
            frete_gratis=float(produto.get("frete_gratis") or 0),
            observacao=produto.get("observacao") or "",
            componentes=componentes_para_json(produto.get("componentes") or []),
        )

        db.add(produto_model)

    db.commit()


def preparar_banco():
    criar_tabelas()

    db = SessionLocal()

    try:
        garantir_empresa_padrao(db)
        migrar_json_antigo_para_banco(db)

    finally:
        db.close()


def listar_produtos(empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    db = SessionLocal()

    try:
        produtos = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .order_by(ProdutoModel.sku.asc())
            .all()
        )

        return [produto_para_dict(produto) for produto in produtos]

    finally:
        db.close()


def carregar_produtos(empresa_id=EMPRESA_PADRAO_ID):
    return listar_produtos(empresa_id=empresa_id)


def salvar_produtos(produtos, empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    db = SessionLocal()

    try:
        db.query(ProdutoModel).filter(ProdutoModel.empresa_id == empresa_id).delete()

        for produto in produtos:
            produto_model = ProdutoModel(
                empresa_id=empresa_id,
                sku=normalizar_sku(produto.get("sku")),
                nome=produto.get("nome") or produto.get("sku"),
                custo=float(produto.get("custo") or 0),
                imposto=float(produto.get("imposto") or 0),
                tipo=produto.get("tipo") or "unitario",
                frete_gratis=float(produto.get("frete_gratis") or 0),
                observacao=produto.get("observacao") or "",
                componentes=componentes_para_json(produto.get("componentes") or []),
                codigos_externos=codigos_para_texto(
                    produto.get("codigosExternos") or produto.get("codigos_externos")
                ),
            )

            db.add(produto_model)

        db.commit()

        return listar_produtos(empresa_id=empresa_id)

    finally:
        db.close()


def buscar_produto_por_sku(sku, empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    sku_normalizado = normalizar_sku(sku)

    db = SessionLocal()

    try:
        produto = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.sku == sku_normalizado)
            .first()
        )

        if not produto:
            return None

        return produto_para_dict(produto)

    finally:
        db.close()


def buscar_produto_por_codigo_externo(codigo, empresa_id=EMPRESA_PADRAO_ID):
    """Acha um produto cujo campo codigos_externos contenha o código informado."""
    preparar_banco()

    codigo_normalizado = normalizar_codigo(codigo)
    if not codigo_normalizado:
        return None

    db = SessionLocal()

    try:
        candidatos = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.codigos_externos.isnot(None))
            .filter(ProdutoModel.codigos_externos != "")
            .filter(ProdutoModel.codigos_externos.like(f"%{codigo_normalizado}%"))
            .all()
        )

        for produto in candidatos:
            if codigo_normalizado in texto_para_codigos(produto.codigos_externos):
                return produto_para_dict(produto)

        return None

    finally:
        db.close()


def cadastrar_produto(dados, empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    db = SessionLocal()

    try:
        sku = normalizar_sku(dados.get("sku"))

        if not sku:
            raise ValueError("SKU é obrigatório.")

        produto_existente = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.sku == sku)
            .first()
        )

        if produto_existente:
            raise ValueError(f"Produto com SKU {sku} já cadastrado.")

        tipo = dados.get("tipo") or "unitario"
        componentes = dados.get("componentes") or []

        custo = float(dados.get("custo") or 0)

        if tipo == "kit_personalizado":
            custo = calcular_custo_kit(db, componentes, empresa_id)

        produto = ProdutoModel(
            empresa_id=empresa_id,
            sku=sku,
            nome=dados.get("nome") or sku,
            custo=custo,
            imposto=float(dados.get("imposto") or 0),
            tipo=tipo,
            frete_gratis=float(dados.get("frete_gratis") or 0),
            observacao=dados.get("observacao") or "",
            componentes=componentes_para_json(componentes),
            codigos_externos=codigos_para_texto(
                dados.get("codigosExternos") or dados.get("codigos_externos")
            ),
        )

        db.add(produto)
        db.commit()
        db.refresh(produto)

        return produto_para_dict(produto)

    finally:
        db.close()


def atualizar_produto(produto_id, dados, empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    db = SessionLocal()

    try:
        produto = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.id == produto_id)
            .first()
        )

        if not produto:
            raise ValueError("Produto não encontrado.")

        sku = normalizar_sku(dados.get("sku"))

        produto_com_mesmo_sku = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.sku == sku)
            .filter(ProdutoModel.id != produto_id)
            .first()
        )

        if produto_com_mesmo_sku:
            raise ValueError(f"Já existe outro produto com SKU {sku}.")

        tipo = dados.get("tipo") or "unitario"
        componentes = dados.get("componentes") or []

        custo = float(dados.get("custo") or 0)

        if tipo == "kit_personalizado":
            custo = calcular_custo_kit(db, componentes, empresa_id)

        produto.sku = sku
        produto.nome = dados.get("nome") or sku
        produto.custo = custo
        produto.imposto = float(dados.get("imposto") or 0)
        produto.tipo = tipo
        produto.frete_gratis = float(dados.get("frete_gratis") or 0)
        produto.observacao = dados.get("observacao") or ""
        produto.componentes = componentes_para_json(componentes)
        if ("codigosExternos" in dados) or ("codigos_externos" in dados):
            produto.codigos_externos = codigos_para_texto(
                dados.get("codigosExternos") or dados.get("codigos_externos")
            )

        db.commit()
        db.refresh(produto)

        return produto_para_dict(produto)

    finally:
        db.close()


def excluir_produto(produto_id, empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    db = SessionLocal()

    try:
        produto = (
            db.query(ProdutoModel)
            .filter(ProdutoModel.empresa_id == empresa_id)
            .filter(ProdutoModel.id == produto_id)
            .first()
        )

        if not produto:
            raise ValueError("Produto não encontrado.")

        db.delete(produto)
        db.commit()

        return {"mensagem": "Produto excluído com sucesso."}

    finally:
        db.close()


def gerar_modelo_importacao_produtos():
    """
    Gera um modelo Excel profissional e didático para importação de produtos.
    Abas: Instruções (guia), Produtos (com exemplos prontos para apagar e usar).
    """
    workbook = Workbook()

    # ---------- Cores / estilos ----------
    azul = "FF1F4E79"
    azul_claro = "FFD6E4F0"
    cinza = "FFF2F2F2"
    branco_negrito = Font(bold=True, color="FFFFFFFF", size=11)
    titulo_font = Font(bold=True, color=azul, size=14)
    negrito = Font(bold=True)
    header_fill = PatternFill("solid", fgColor=azul)
    exemplo_fill = PatternFill("solid", fgColor=cinza)
    wrap = Alignment(wrap_text=True, vertical="top")

    # ============================================================
    # ABA 1 — INSTRUÇÕES (abre primeiro)
    # ============================================================
    guia = workbook.active
    guia.title = "Instruções"
    guia.sheet_view.showGridLines = False

    guia["A1"] = "Como importar seus produtos"
    guia["A1"].font = titulo_font

    linhas_guia = [
        ("", ""),
        ("Passo a passo", ""),
        ("1.", "Vá na aba 'Produtos' (na parte de baixo)."),
        ("2.", "Apague as linhas de exemplo e preencha com os seus produtos."),
        ("3.", "Salve o arquivo e envie em 'Importar produtos' no sistema."),
        ("4.", "SKUs que já existem são ATUALIZADOS; novos são criados."),
        ("", ""),
        ("As colunas", ""),
        ("tipo", "unitario  ou  kit_personalizado"),
        ("sku", "O código do produto (ou do kit). Ex.: CV017, ANDELEVA."),
        ("nome", "Nome do produto como você quer ver no sistema."),
        (
            "custo_unitario",
            "Quanto VOCÊ paga pelo produto (seu custo). Use ponto: 12.50",
        ),
        ("componente_sku", "Só para kit: o SKU de cada item que compõe o kit."),
        (
            "componente_quantidade",
            "Só para kit: quantas unidades daquele item entram no kit.",
        ),
        ("observacao", "Campo livre (anotações suas)."),
        (
            "codigo_marketplace",
            "ID do produto no marketplace (ex.: ID do produto na Shopee). Liga o relatório ao seu SKU.",
        ),
        ("", ""),
        ("1) Produto simples (unitário)", ""),
        (
            "",
            "tipo = unitario, preencha sku, nome e custo_unitario. Deixe as colunas de",
        ),
        ("", "componente em branco. É o caso da maioria dos produtos."),
        ("", ""),
        ("2) Variações (cor/tamanho)", ""),
        ("", "Cada variação é um produto unitário com SEU PRÓPRIO SKU."),
        (
            "",
            "Ex.: CAMISETA-AZUL-P, CAMISETA-AZUL-M, CAMISETA-PRETA-P — uma linha cada,",
        ),
        ("", "todas como 'unitario', com o custo de cada uma."),
        ("", ""),
        ("3) Kit montado por você (kit_personalizado)", ""),
        ("", "Use quando o kit é formado por outros produtos que você já cadastra."),
        ("", "Repita o SKU do kit em VÁRIAS linhas, uma por componente, preenchendo"),
        ("", "componente_sku e componente_quantidade. O custo do kit é calculado"),
        (
            "",
            "automaticamente somando o custo dos componentes (não preencha custo_unitario).",
        ),
        ("", ""),
        ("4) Kit automático por sufixo -KITn (não precisa cadastrar!)", ""),
        ("", "Se o anúncio vende N unidades do MESMO produto, basta o SKU terminar em"),
        ("", "-KIT2, -KIT3 etc. Ex.: cadastre SACO5060 (custo de 1 unidade); quando o"),
        (
            "",
            "pedido vier como SACO5060-KIT2, o sistema multiplica o custo por 2 sozinho.",
        ),
    ]
    for rotulo, texto in linhas_guia:
        guia.append([rotulo, texto])

    for row in guia.iter_rows(min_row=2, max_row=guia.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value in (
                "Passo a passo",
                "As colunas",
                "1) Produto simples (unitário)",
                "2) Variações (cor/tamanho)",
                "3) Kit montado por você (kit_personalizado)",
                "4) Kit automático por sufixo -KITn (não precisa cadastrar!)",
            ):
                cell.font = negrito
    guia.column_dimensions["A"].width = 24
    guia.column_dimensions["B"].width = 86

    # ============================================================
    # ABA 2 — PRODUTOS (cabeçalho + exemplos)
    # ============================================================
    sheet = workbook.create_sheet("Produtos")
    colunas = [
        "tipo",
        "sku",
        "nome",
        "custo_unitario",
        "componente_sku",
        "componente_quantidade",
        "observacao",
        "codigo_marketplace",
    ]
    sheet.append(colunas)
    for c in range(1, len(colunas) + 1):
        cell = sheet.cell(row=1, column=c)
        cell.font = branco_negrito
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"

    exemplos = [
        [
            "unitario",
            "CV017",
            "Rechaud Inox 11L 2 Cubas",
            75.0,
            "",
            "",
            "Produto simples",
            "",
        ],
        [
            "unitario",
            "ANDELEVA",
            "Andador Idoso Dobrável Alumínio",
            120.0,
            "",
            "",
            "Produto simples",
            "",
        ],
        [
            "unitario",
            "POTE525ML",
            "Kit 10 Potes Herméticos 525ml",
            28.0,
            "",
            "",
            "ID da Shopee em codigo_marketplace",
            "58200017353",
        ],
        [
            "unitario",
            "CAMISETA-AZUL-P",
            "Camiseta Azul Tam P",
            18.9,
            "",
            "",
            "Variação cor/tamanho",
            "",
        ],
        [
            "kit_personalizado",
            "KIT-SACO-BOMBA",
            "Kit Saco a Vácuo + Bomba",
            "",
            "SACO5060",
            2,
            "Kit: 1 linha por componente",
            "",
        ],
        [
            "kit_personalizado",
            "KIT-SACO-BOMBA",
            "Kit Saco a Vácuo + Bomba",
            "",
            "TOP0864-2",
            1,
            "Mesmo SKU do kit, 2º componente",
            "",
        ],
    ]
    for linha in exemplos:
        sheet.append(linha)
    for r in range(2, 2 + len(exemplos)):
        for c in range(1, len(colunas) + 1):
            sheet.cell(row=r, column=c).fill = exemplo_fill

    larguras = [18, 18, 36, 14, 18, 22, 34, 20]
    for i, w in enumerate(larguras, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w

    nota = sheet.cell(row=2 + len(exemplos) + 1, column=1)
    nota.value = "↑ Apague as linhas de exemplo acima e preencha com os seus produtos."
    nota.font = Font(italic=True, color="FF888888")

    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)

    return arquivo


def _ler_excel_tolerante(caminho_arquivo, **_kwargs):
    """Compatibilidade: devolve lista de dicts. Usa openpyxl, sem pandas."""
    return _xlsx_para_lista(caminho_arquivo)


def importar_mapa_shopee(caminho_arquivo, empresa_id=EMPRESA_PADRAO_ID):
    """Importa o mapa de produtos da Shopee (mass_update_basic_info)."""
    preparar_banco()

    wb = _abrir_xlsx(caminho_arquivo)
    ws = wb.active
    todas = list(ws.iter_rows(values_only=True))
    wb.close()

    linha_cab = None
    for i, row in enumerate(todas[:10]):
        if any("id do produto" in str(v).strip().lower() for v in row if v):
            linha_cab = i
            break
    if linha_cab is None:
        raise ValueError(
            "Planilha da Shopee não reconhecida (não encontrei 'ID do Produto')."
        )

    cab = [
        str(h).strip() if h is not None else f"_col{i}"
        for i, h in enumerate(todas[linha_cab])
    ]
    linhas_dados = [{c: v for c, v in zip(cab, row)} for row in todas[linha_cab + 1 :]]

    def achar(opcoes):
        for c in cab:
            if any(o in str(c).strip().lower() for o in opcoes):
                return c
        return None

    col_id = achar(["id do produto"])
    col_sku = achar(["sku de referência", "sku de referencia", "sku"])
    col_nome = achar(["nome do produto", "nome"])

    if not col_id or not col_sku:
        raise ValueError("Não encontrei as colunas de ID do Produto e SKU.")

    vinculados, erros = 0, []

    for linha in linhas_dados:
        try:
            id_produto = normalizar_codigo(linha.get(col_id))
            sku = normalizar_sku(linha.get(col_sku))
            nome = str(linha.get(col_nome) or sku).strip() if col_nome else sku
            if not id_produto or not id_produto.isdigit() or not sku:
                continue
            existente = buscar_produto_por_sku(sku, empresa_id=empresa_id)
            if existente:
                codigos = set(texto_para_codigos(existente.get("codigosExternos")))
                codigos.add(id_produto)
                atualizar_produto(
                    existente["id"],
                    {
                        "sku": sku,
                        "nome": existente.get("nome") or nome,
                        "custo": existente.get("custo") or 0,
                        "imposto": 0,
                        "tipo": existente.get("tipo") or "unitario",
                        "frete_gratis": 0,
                        "observacao": existente.get("observacao") or "",
                        "componentes": existente.get("componentes") or [],
                        "codigosExternos": ",".join(sorted(codigos)),
                    },
                    empresa_id=empresa_id,
                )
            else:
                cadastrar_produto(
                    {
                        "sku": sku,
                        "nome": nome,
                        "custo": 0,
                        "tipo": "unitario",
                        "codigosExternos": id_produto,
                        "observacao": "Importado do mapa da Shopee (defina o custo).",
                    },
                    empresa_id=empresa_id,
                )
            vinculados += 1
        except Exception as e:
            erros.append({"sku": str(linha.get(col_sku) or ""), "erro": str(e)})

    return {
        "total_vinculados": vinculados,
        "total_erros": len(erros),
        "erros": erros,
        "observacao": "Produtos vinculados. Defina o custo dos novos (entraram com 0).",
    }


def importar_produtos_excel(caminho_arquivo, empresa_id=EMPRESA_PADRAO_ID):
    preparar_banco()

    linhas_dados = _xlsx_para_lista(caminho_arquivo)

    criados = 0
    atualizados = 0
    erros = []
    produtos_agrupados = {}

    for idx, linha in enumerate(linhas_dados):
        linha_excel = idx + 2
        try:
            tipo = str(linha.get("tipo") or "unitario").strip()
            sku = normalizar_sku(linha.get("sku"))
            nome = str(linha.get("nome") or sku or "").strip()
            custo_unitario = linha.get("custo_unitario")
            componente_sku = normalizar_sku(linha.get("componente_sku"))
            componente_quantidade = linha.get("componente_quantidade")
            observacao = str(linha.get("observacao") or "").strip()
            codigo_marketplace = linha.get("codigo_marketplace") or linha.get(
                "id_produto"
            )

            if not sku:
                continue

            if sku not in produtos_agrupados:
                custo = 0.0
                try:
                    if custo_unitario is not None:
                        custo = float(custo_unitario)
                except (TypeError, ValueError):
                    pass

                cod_ext = ""
                if codigo_marketplace is not None:
                    cod_ext = normalizar_codigo(codigo_marketplace)

                produtos_agrupados[sku] = {
                    "linha_excel": linha_excel,
                    "tipo": tipo,
                    "sku": sku,
                    "nome": nome,
                    "custo": custo,
                    "imposto": 0,
                    "frete_gratis": 0,
                    "observacao": observacao,
                    "componentes": [],
                    "codigosExternos": cod_ext,
                }

            if tipo == "kit_personalizado" and componente_sku:
                qtd = 1.0
                try:
                    if componente_quantidade is not None:
                        qtd = float(componente_quantidade)
                except (TypeError, ValueError):
                    pass
                produtos_agrupados[sku]["componentes"].append(
                    {"sku": componente_sku, "quantidade": qtd}
                )

        except Exception as erro:
            erros.append(
                {
                    "linha_excel": linha_excel,
                    "sku": str(linha.get("sku") or ""),
                    "erro": str(erro),
                }
            )

    for sku, produto in produtos_agrupados.items():
        try:
            existente = buscar_produto_por_sku(sku, empresa_id=empresa_id)

            if existente:
                atualizar_produto(
                    existente["id"],
                    produto,
                    empresa_id=empresa_id,
                )
                atualizados += 1
            else:
                cadastrar_produto(produto, empresa_id=empresa_id)
                criados += 1

        except Exception as erro:
            erros.append(
                {
                    "linha_excel": produto.get("linha_excel"),
                    "sku": sku,
                    "erro": str(erro),
                }
            )

    return {
        "total_criados": criados,
        "total_atualizados": atualizados,
        "total_erros": len(erros),
        "erros": erros,
    }


preparar_banco()
